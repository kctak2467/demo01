import csv
import io
import urllib
import openpyxl
from openpyxl.styles import Font
from openpyxl.styles.alignment import Alignment
from openpyxl.styles import PatternFill
from openpyxl.styles.borders import Border, Side
from openpyxl.worksheet.page import PageMargins,PrintPageSetup
from openpyxl.worksheet.worksheet import Worksheet

from django.shortcuts import render
from django.views.generic import ListView,DetailView,CreateView,UpdateView,DeleteView
from django.urls import reverse, reverse_lazy
from django.db.models import Q
from django.contrib import messages
from django.views import generic
from django.http import HttpResponse

from .models import Kiki
from .forms import SearchForm
from .forms import CSVUploadForm
from .forms import ExcelUploadForm

# 機器台帳（検索・一覧）
class KikiList(ListView):
    template_name = 'kiki/kiki_list.html'
    # ListViewにおけるテンプレート名(HTML)はデフォルトで[object名]_List.htmlとなるため
    # 同名であれば、template_nameの指定は不要
    context_object_name = 'kiki_list'
    # テンプレート名(HTML)内のオブジェクトデフォルト名（object_list）を別名で利用する場合は
    # context_object_nameを定義する。
    model = Kiki
    # 表示件数指定は「paginate_by = 3」を追加
    # QuerySet（クエリセット）を利用する場合は model = kikiを以下のように置き換える
    #    queryset = Member.objects.all()     全件
    #    queryset = Member.objects.order_by('-age')    降順
    #    queryset = Member.objects.filter(juyodo='A')  フィルタリング

    # --- 検索画面追加バージョン
    # paginate_by = 5

    def post(self, request, *args, **kwargs):
        form_value = [
            self.request.POST.get('kiki_id', None),
            self.request.POST.get('kiki_name', None),
        ]
        request.session['form_value'] = form_value
        # 検索時にページネーションに関連したエラーを防ぐ
        self.request.GET = self.request.GET.copy()
        self.request.GET.clear()
        return self.get(request, *args, **kwargs)

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        # sessionに値がある場合、その値をセットする。（ページングしてもform値が変わらないように）
        kiki_id = ''
        kiki_name = ''
        if 'form_value' in self.request.session:
            form_value = self.request.session['form_value']
            kiki_id = form_value[0]
            kiki_name = form_value[1]
        default_data = {'kiki_id': kiki_id,  # タイトル
                        'kiki_name': kiki_name,  # 内容
                        }
        test_form = SearchForm(initial=default_data) # 検索フォーム
        context['test_form'] = test_form
        return context

    def get_queryset(self):
        # sessionに値がある場合、その値でクエリ発行する。
        if 'form_value' in self.request.session:
            form_value = self.request.session['form_value']
            kiki_id = form_value[0]
            kiki_name = form_value[1]

            # 検索条件
            condition_kiki_id = Q()
            condition_kiki_name = Q()
            if len(kiki_id) != 0 and kiki_id[0]:
                condition_kiki_id = Q(kiki_id__contains=kiki_id)
            if len(kiki_name) != 0 and kiki_name[0]:
                condition_kiki_name = Q(kiki_name__contains=kiki_name)

            return Kiki.objects.select_related().filter(condition_kiki_id & condition_kiki_name)
        else:
            # 何も返さない
            return Kiki.objects.none()

# 機器台帳（参照）
class KikiDetail(DetailView):
    # template_name = 'kiki/kiki_detail.html'
    model = Kiki
    # def get_context_data(self, **kwargs):
    #     context = super().get_context_data(**kwargs)
    #     messages.info(self.request, 'テスト')
    #     return context

# 機器台帳（登録）
class KikiCreate(CreateView):
    # template_name = 'kiki/kiki_form.html'
    model = Kiki
    fields = ['kiki_id', 'kiki_name', 'keito', 'settibasho', 'juyodo', 'nensu']

    def get_success_url(self):
        messages.info(self.request, '登録しました。')
        return reverse('detail', kwargs={'pk': self.object.pk})

    def get_form(self):
        form = super(KikiCreate, self).get_form()
#        form.fields['kiki_name'].label = '機器名称'   #ラベル
#        form.initial['kiki_name'] = '機器名称ですよ'  #初期値設定
#       form.fields['settibasho'].required = False  #必須項目OFF
        return form

# 機器台帳（更新）
class KikiUpdate(UpdateView):
    template_name = 'kiki/kiki_update_form.html'
    model = Kiki
    fields = ['kiki_id', 'kiki_name', 'keito', 'settibasho', 'juyodo', 'nensu']


    # def get_context_data(self, **kwargs):
    #     context = super().get_context_data(**kwargs)
    #
    #     context['title'] = '編集画面（入力）'
    #     context['mode'] = 'input'
    #     return context

    def get_success_url(self):
        messages.info(self.request, '更新しました。')
        return reverse('detail', kwargs={'pk': self.object.pk})

    def get_form(self):
        form = super(KikiUpdate, self).get_form()
#        form.fields['kiki_name'].label = '機器名称'   #ラベル
#       form.initial['kiki_name'] = '機器名称ですよ'   #初期値設定
#        form.fields['settibasho'].required = False  #必須項目OFF
        return form

# 機器台帳（削除）
class KikiDelete(DeleteView):
    # template_name = 'kiki/kiki_confirm_delete.html'
    model = Kiki

    # success_url = reverse_lazy('kiki')

    def get_success_url(self):
        messages.info(self.request, '削除しました。')
        return reverse('kiki')

# 機器台帳(csvアップロード)
class KikiImport(generic.FormView):

    template_name = 'kiki/kiki_import.html'
    success_url = reverse_lazy('kiki')
    form_class = CSVUploadForm

    def get_context_data(self, **kwargs):
        ctx = super().get_context_data(**kwargs)
        ctx['form_name'] = 'csvdownload'
        # ctx['filename'] = self.request.FILES['filename'].name
        return ctx

    def form_valid(self, form):
        # """postされたCSVファイルを読み込み、役職テーブルに登録します"""
        #csvfile = form.cleaned_data['file']
        csvfile = io.TextIOWrapper(form.cleaned_data['file'])
        reader = csv.reader(csvfile)
        for row in reader:
            # """
            # 役職テーブルを役職コード(primary key)で検索します
            # """
            kiki, created = Kiki.objects.get_or_create(pk=row[0])
            kiki.kiki_id = row[1]
            kiki.kiki_name = row[2]
            kiki.keito = row[3]
            kiki.settibasho = row[4]
            kiki.juyodo = row[5]
            kiki.nensu = row[6]
            kiki.save()
        return super().form_valid(form)

    def get_success_url(self):
        messages.info(self.request, 'アップロードしました。')
        return reverse('import')

# 機器台帳（csvダウンロード）
def KikiExport(request):

    response = HttpResponse(content_type='text/csv; charset=Shift-JIS')
    filename = urllib.parse.quote((u'kiki_download.csv').encode("utf8"))
    response['Content-Disposition'] = 'attachment; filename*=UTF-8\'\'{}'.format(filename)
    writer = csv.writer(response)

    if 'form_value' in request.session:
        form_value = request.session['form_value']
        kiki_id = form_value[0]
        kiki_name = form_value[1]

        # 検索条件
        condition_kiki_id = Q()
        condition_kiki_name = Q()
        if len(kiki_id) != 0 and kiki_id[0]:
            condition_kiki_id = Q(kiki_id__contains=kiki_id)
        if len(kiki_name) != 0 and kiki_name[0]:
            condition_kiki_name = Q(kiki_name__contains=kiki_name)

    for kiki in Kiki.objects.select_related().filter(condition_kiki_id & condition_kiki_name):
        writer.writerow([kiki.pk, kiki.kiki_id , kiki.kiki_name , kiki.keito , kiki.settibasho , kiki.juyodo , kiki.nensu ])

    return response

# 機器台帳（excelダウンロード）
def KikiDownloadExcel(request):

    # content-type of response
    response = HttpResponse(content_type='application/ms-excel')
    # decide file name
    response['Content-Disposition'] = 'attachment; filename="kiki_download.xlsx"'

    # Excelのテンプレートファイルの読み込み
    wb = openpyxl.load_workbook('kiki/exceltemplates/kiki_sample.xlsx')

    # adding sheet
    ws= wb.active

    # Sheet header, first row

    #ページ余白（インチ設定（20mm:0.78704016、15mm=0.5905512）
    # pagemargins =PageMargins(top=0.7874016,bottom=0.7874016,left=0.5905512,right=0.5905512,
    #                          header=0.5905512,footer=0.5905512)
    # # ページ設定
    # ws.print_title_rows = "1:3"                 #行タイトル範囲
    # ws.paper_size = ws.PAPERSIZE_A4                  #用紙サイズ
    #
    # ws.page_margins = pagemargins
    # fill =PatternFill(patternType='solid', fgColor='ccffff')
    # column header names, you can use your own headers here
    # columns = ['キー', '機器ID', '機器名称', '系統', '設置場所', '重要度', '耐用年数', ]
    # columns_width = [6, 20, 20, 15 , 15, 10, 10, ]

    #行の高さ
    # row_height = 15
    # ws.row_dimensions[1].height = row_height

    # # write column headers in sheet
    # for col_num in range(len(columns)):
    #
    #     #ヘッダタイトルの設定
    #     ws.cell(row=row_num, column=col_num + 1).value = columns[col_num]
    #     ws.cell(row=row_num, column=col_num + 1).font = font
    #     ws.cell(row=row_num, column=col_num + 1).alignment = alignment
    #     ws.cell(row=row_num, column=col_num + 1).fill = fill
    #     ws.cell(row=row_num, column=col_num + 1).border = border
    #
    #     # 列幅の設定（列数(1列目)から列番号(A)を取得し、列幅を設定）
    #     col_name = openpyxl.utils.get_column_letter(col_num + 1)
    #     ws.column_dimensions[col_name].width = columns_width[col_num]

    row_height = 15
    font = Font(size=9, name="Meiryo UI")
    alignment = Alignment(horizontal='center', vertical='center', wrap_text=False)
    border = Border(top=Side(style='thin', color='000000'),
                    bottom=Side(style='thin', color='000000'),
                    left=Side(style='thin', color='000000'),
                    right=Side(style='thin', color='000000'))

    if 'form_value' in request.session:
        form_value = request.session['form_value']
        kiki_id = form_value[0]
        kiki_name = form_value[1]

        # 検索条件
        condition_kiki_id = Q()
        condition_kiki_name = Q()
        if len(kiki_id) != 0 and kiki_id[0]:
            condition_kiki_id = Q(kiki_id__contains=kiki_id)
        if len(kiki_name) != 0 and kiki_name[0]:
            condition_kiki_name = Q(kiki_name__contains=kiki_name)

        # get your data, from database or from a text file...
        # data = get_data()  # dummy method to fetch data.

        cnt = 0  # データ件数
        for data in Kiki.objects.select_related().filter(condition_kiki_id & condition_kiki_name):
            cnt = cnt + 1
            row_num = cnt + 3
            columns = [cnt, data.pk, data.kiki_id, data.kiki_name, data.keito, data.settibasho, data.juyodo, data.nensu, ]
            for col_num in range(len(columns)):
                ws.cell(row=row_num, column=col_num + 1).value = columns[col_num]
                ws.cell(row=row_num, column=col_num + 1).font = font
                ws.cell(row=row_num, column=col_num + 1).border = border
                ws.row_dimensions[row_num].height = row_height
                if col_num == 0 or col_num == 1 or col_num == 6 or col_num == 7:
                    ws.cell(row=row_num, column=col_num + 1).alignment = alignment

        wb.save(response)
        return response

# 機器台帳(Excelアップロード)




class KikiUploadExcel(generic.FormView):

    template_name = 'kiki/kiki_upload_excel.html'
    success_url = reverse_lazy('kiki')
    form_class = ExcelUploadForm

    def get_context_data(self, **kwargs):
        ctx = super().get_context_data(**kwargs)
        ctx['form_name'] = 'exceldownload'
        return ctx

    def form_valid(self, form):

        excel_file = self.request.FILES["file"]

        if 'form_value' in self.request.session:
            form_value = self.request.session['form_value']

        # excel_file = form_value.cleaned_data['file1']
        # you may put validations here to check extension or file size

        wb = openpyxl.load_workbook(excel_file)

        # getting a particular sheet by name out of many sheets
        worksheet = wb["Sheet1"]
        print(worksheet)

        excel_data = list()

        # 行を繰り返し処理し、行の各セルから値を取得
        cnt = 0
        for row in worksheet.iter_rows():
            cnt = cnt + 1
            if cnt >= 4:
                row_data = list()
                for cell in row:
                    row_data.append(str(cell.value))
                    print(cell.value)
                excel_data.append(row_data)

        for row in excel_data:
            kiki, created = Kiki.objects.get_or_create(pk=row[1])
            kiki.kiki_id = row[2]
            kiki.kiki_name = row[3]
            kiki.keito = row[4]
            kiki.settibasho = row[5]
            kiki.juyodo = row[6]
            kiki.nensu = row[7]
            kiki.save()

        return super().form_valid(form)
#
    def get_success_url(self):
        messages.info(self.request, 'アップロードしました。')
        return reverse('kiki')

