from django.shortcuts import render
import openpyxl

def index(request):
    if "GET" == request.method:
        return render(request, 'excelupload/index.html', {})
    else:
        excel_file = request.FILES["excel_file"]

        # ここにバリデーショションを配置して、拡張子またはファイルサイズを確認できる

        wb = openpyxl.load_workbook(excel_file)

        # # 全てのシートを取得
        # sheets = wb.sheetnames
        # print(sheets)

        # シート名を指定して取得
        worksheet = wb["Sheet1"]
        print(worksheet)

        # # アクティブシートを取得
        # active_sheet = wb.active
        # print(active_sheet)

        # # セルの取得
        # print(worksheet["A1"].value)

        excel_data = list()

        # 行を繰り返し処理し、行の各セルから値を取得
        for row in worksheet.iter_rows():
            row_data = list()
            for cell in row:
                row_data.append(str(cell.value))
                print(cell.value)
            excel_data.append(row_data)

        return render(request, 'excelupload/index.html', {"excel_data":excel_data})